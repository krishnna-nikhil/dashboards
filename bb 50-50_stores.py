#!/usr/bin/env python
# coding: utf-8

# In[3]:


import pandas as pd

# Path to the Excel file
file_path = "/Users/devnikhil/Downloads/Sales Analysis Report (sale.report) - 2025-06-28T175844.229.xlsx"

# Load the Excel file into a DataFrame
df = pd.read_excel(file_path)

# Create the 'fifty_disc' column based on the condition for 'Discount %'

df['Total'].fillna(0, inplace=True)




df.info()

# Display unique companies
unique_companies = df['Company'].unique()

# Print the unique companies
print(unique_companies)

# Filter out 'BRYT BAZAAR -BLR' and 'BRYT BAZAAR RICHMOND TOWN'
df = df[~df['Company'].isin(['BRYT BAZAAR -BLR','Bryt Bazaar Indiranagar', 'BRYT BAZAAR RICHMOND TOWN','SRI DEEPA RETAIL','DEEPA RETAIL','BRYT BAZAAR YELLAHANKA'])]

# Rename specific company names
df['Company'] = df['Company'].replace({
    'BRYT BAZAAR J P NAGAR': 'JPN',
    'BRYT BAZAAR ITI': 'ITI',
    'BRYT BAZAAR KUMARSWAMY LAYOUT': 'KSL',
    'BRYT BAZAAR BASAPURA MAIN ROAD': 'BMR',
    'BRYT BAZAAR RAJAJINAGAR': 'RJJ',
    'BRYT BAZAAR RK HEGDE NAGAR': 'RKH',
    'Bryt Bazaar Indiranagar': 'IND'
})

# Optional: Print to verify the changes
print(df['Company'].unique())
df = df[df['Customer'] != 'SRI DEEPA RETAIL']

df = df[df['Customer'] != 'KIKO']





print("Specified products have been removed from the DataFrame.")

df.info()


# In[146]:


import pandas as pd

# Ensure 'Order Date' is in datetime format
df['Order Date'] = pd.to_datetime(df['Order Date'])

# Split 'Order Date' into separate 'date' and 'time' columns
df['date'] = df['Order Date'].dt.date

# Determine the earliest date
earliest_date = df['date'].min()

# Calculate the 'week' column starting from the earliest date
df['week'] = ((pd.to_datetime(df['date']) - pd.to_datetime(earliest_date)).dt.days // 7) + 1

# Display the updated DataFrame
df[['date', 'week']].drop_duplicates().sort_values(by='date')


# In[147]:


# Step 1: Get unique bills per Order Date
bill_summary = df.groupby('Order Date', as_index=False)['Total'].sum()

# Step 2: Merge to get week info
week_map = df[['Order Date', 'week']].drop_duplicates()
bill_summary = bill_summary.merge(week_map, on='Order Date', how='left')

# Step 3: Define proper range buckets
def classify_range(total):
    if total < 500:
        return '<500'
    elif 500 <= total < 1000:
        return '500-999'
    elif 1000 <= total < 2000:
        return '1000-1999'
    elif 2000 <= total < 3000:
        return '2000-2999'
    else:
        return '3000+'

bill_summary['Range'] = bill_summary['Total'].apply(classify_range)

# Step 4: Group by week and range
weekly_summary = (
    bill_summary.groupby(['week', 'Range'])
    .size()
    .unstack(fill_value=0)
    .reset_index()
)

# Step 5: Ensure column order
col_order = ['week', '<500', '500-999', '1000-1999', '2000-2999', '3000+']
for col in col_order:
    if col not in weekly_summary.columns:
        weekly_summary[col] = 0
weekly_summary = weekly_summary[col_order]

# Output
print(weekly_summary)


# In[148]:


# Step 1: Get unique bills per Order Date
bill_summary = df.groupby('Order Date', as_index=False)['Total'].sum()

# Step 2: Merge week info
week_map = df[['Order Date', 'week']].drop_duplicates()
bill_summary = bill_summary.merge(week_map, on='Order Date', how='left')

# Step 3: Filter for week 4 and Total > 3000
week_4_above_3000 = bill_summary[(bill_summary['week'] == 4) & (bill_summary['Total'] >= 3000)]

# Step 4: Display unique order dates
print("Unique Order Dates with bills above 3000 in week 4:")
print(week_4_above_3000[['Order Date', 'Total']].sort_values(by='Total', ascending=False))


# In[ ]:





# In[ ]:





# In[150]:


import pandas as pd

# Filter data for Week 4
week_4_data = df[df['week'] == 4]

# Find the most-selling SKU for Week 4
most_selling_sku = week_4_data.groupby('Product Variant')['Total'].sum().idxmax()
most_selling_total = week_4_data.groupby('Product Variant')['Total'].sum().max()

# Calculate total quantity ordered for the most-selling SKU
most_selling_sku_quantity = week_4_data[week_4_data['Product Variant'] == most_selling_sku]['Qty Ordered'].sum()

# Format the most-selling SKU data into a DataFrame
most_selling_sku_data = pd.DataFrame({
    'Product Variant': [most_selling_sku],
    'Total Sales': [most_selling_total],
    'Total Quantity Ordered': [most_selling_sku_quantity]
})

# Calculate total sales by company and format into a DataFrame
branch_total_sales = week_4_data.groupby('Company')['Total'].sum().sort_values(ascending=False)
branch_total_sales_df = branch_total_sales.reset_index()
branch_total_sales_df.columns = ['Company', 'Sales']

# Combine and display results
print("top sku")
print(most_selling_sku_data.to_string(index=False))

print("total sales")
print(branch_total_sales_df.to_string(index=False))


# In[151]:


comparison_df = df[df['week'].isin([3, 4])].groupby(['Company', 'week'])['Total'].sum().unstack().assign(
    Difference=lambda x: x[4] - x[3],
    Percentage_Change=lambda x: (x[4] - x[3]) / x[3] * 100
).reset_index().rename(columns={3: 'Week 3', 4: 'Week 4', 'Change': 'Difference (W4-W3)', 'Change %': 'Percentage Change (%)'})
# Reset index and reorder based on a custom list of company names
custom_order = ['JPN', 'KSL', 'BMR', 'ITI', 'RKH', 'RJJ']
comparison_df = comparison_df.set_index('Company').reindex(custom_order).reset_index()

# Display the final DataFrame
print(comparison_df)


# In[152]:


import pandas as pd

# Assuming 'df' is your DataFrame that contains the sales data
# Example: df = pd.read_csv('your_data.csv')

# Initialize an empty list to store summary rows
weekly_summary_data = []

# Define the column names for the summary DataFrame
columns = [
    "Company",
    "Total Sales W1", "Total Sales W2", "Total Sales W3", "Total Sales W4",
    "AOV W1", "AOV W2", "AOV W3", "AOV W4",
    "NOB W1", "NOB W2", "NOB W3", "NOB W4",
    "ADS W1", "ADS W2", "ADS W3", "ADS W4"  # Added Average Daily Sales (ADS)
]

# Iterate over each unique company in the dataset
for company in df['Company'].unique():
    # Create a list to hold data for each company
    company_data = [company]
    
    # Calculate Total Sales, AOV, and NOB for each week
    total_sales = []
    aov = []
    nob = []
    ads = []  # For Average Daily Sales (ADS)
    
    for week in range(1, 5):  # Iterate over weeks 1 to 4
        # Filter the data for the current company and week
        weekly_data = df[(df['Company'] == company) & (df['week'] == week)]
        
        # Calculate Total Sales for the current week
        week_sales = weekly_data['Total'].sum()
        total_sales.append(week_sales)
        
        # Calculate NOB (Number of Unique Order Dates) for the current week
        unique_order_dates = weekly_data['Order Date'].nunique()
        nob.append(unique_order_dates)
        
        # Calculate AOV for the current week
        week_aov = week_sales / unique_order_dates if unique_order_dates > 0 else 0
        aov.append(week_aov)
        
        # Calculate Average Daily Sales (ADS) for the current week
        week_ads = week_sales / 7  # Assuming 7 days in a week
        ads.append(week_ads)
    
    # Append the calculated data for the current company to the list
    company_data.extend(total_sales + aov + nob + ads)
    
    # Append the company data to the summary list
    weekly_summary_data.append(company_data)

# Convert the summary data to a DataFrame
weekly_summary_df = pd.DataFrame(weekly_summary_data, columns=columns)

# Reorder the rows as per the desired order: 'JPN', 'KSL', 'BMR', 'ITI', 'RJJ', 'RKH'
ordered_companies = ['JPN', 'KSL', 'BMR', 'ITI', 'RKH', 'RJJ']
weekly_summary_df = weekly_summary_df.set_index('Company').loc[ordered_companies].reset_index()

# Print the final DataFrame
print(weekly_summary_df)


# In[153]:


# Step 1: Filter the DataFrame for week == 4
df_week_4 = df[df['week'] == 4]

# Step 2: Create the 'fifty_disc' column to categorize whether the discount is 50% or more
df_week_4['fifty_disc'] = df_week_4['Discount %'].apply(lambda x: 'yes' if x >= 50 else 'no')

# Step 3: Group by company and calculate the total sales, total normal sales, and total 50% discount sales for week 4
company_sales_summary_week_4 = df_week_4.groupby('Company').apply(lambda group: pd.Series({
    'Total Sales': group['Total'].sum(),
    'Total Normal Sales': group[group['fifty_disc'] == 'no']['Total'].sum(),
    'Total 50% Discount Sales': group[group['fifty_disc'] == 'yes']['Total'].sum()
})).reset_index()

# Step 4: Calculate the percentages for normal and 50% discount sales
company_sales_summary_week_4['Total Normal Sales %'] = (
    company_sales_summary_week_4['Total Normal Sales'] / company_sales_summary_week_4['Total Sales'] * 100
)

company_sales_summary_week_4['Total 50% Discount Sales %'] = (
    company_sales_summary_week_4['Total 50% Discount Sales'] / company_sales_summary_week_4['Total Sales'] * 100
)

# Step 5: Reorder the companies as per the specified order: 'JPN', 'KSL', 'BMR', 'ITI', 'RKH', 'RJJ'
ordered_companies = ['JPN', 'KSL', 'BMR', 'ITI', 'RKH', 'RJJ']
company_sales_summary_week_4 = company_sales_summary_week_4.set_index('Company').loc[ordered_companies].reset_index()

# Print the updated DataFrame with the percentages and the ordered companies
print(company_sales_summary_week_4)


# In[154]:


# Step 1: Filter the DataFrame for week == 4
df_week_4 = df[df['week'] == 4]

# Step 2: Create the 'fifty_disc' column to categorize whether the discount is 50% or more
df_week_4['fifty_disc'] = df_week_4['Discount %'].apply(lambda x: 'yes' if x >= 50 else 'no')

# Step 3: Group by company and calculate the total sales, total normal sales, and total 50% discount sales for week 4
company_sales_summary_week_4 = df_week_4.groupby('Company').apply(lambda group: pd.Series({
    'Total Sales': group['Total'].sum(),
    'Total Normal Sales': group[group['fifty_disc'] == 'no']['Total'].sum(),
    'Total 50% Discount Sales': group[group['fifty_disc'] == 'yes']['Total'].sum()
})).reset_index()

# Step 4: Calculate the percentages for normal and 50% discount sales
company_sales_summary_week_4['Total Normal Sales %'] = (
    company_sales_summary_week_4['Total Normal Sales'] / company_sales_summary_week_4['Total Sales'] * 100
)

company_sales_summary_week_4['Total 50% Discount Sales %'] = (
    company_sales_summary_week_4['Total 50% Discount Sales'] / company_sales_summary_week_4['Total Sales'] * 100
)

# Step 5: Print the result as is (no reordering)
print(company_sales_summary_week_4)


# In[155]:


import pandas as pd
import matplotlib.pyplot as plt

# Filter data for Week 1 and Week 2
week_1_data = df[df['week'] == 3]
week_2_data = df[df['week'] == 4]

# Aggregate sales data by 'Product Variant' for Week 1
week_1_aggregated = week_1_data.groupby('Product Variant', as_index=False)['Total'].sum()

# Aggregate sales data by 'Product Variant' for Week 2
week_2_aggregated = week_2_data.groupby('Product Variant', as_index=False)['Total'].sum()

# Merge the aggregated data for Week 1 and Week 2
comparison_data = pd.merge(week_1_aggregated, 
                           week_2_aggregated, 
                           on='Product Variant', 
                           how='outer', 
                           suffixes=('_Week1', '_Week2'))

# Fill NaN values with 0 for products that might be missing in either week
comparison_data.fillna(0, inplace=True)

# Calculate the revenue loss between Week 1 and Week 2 (Sales in Week 1 - Sales in Week 2)
comparison_data['Revenue Loss'] = comparison_data['Total_Week1'] - comparison_data['Total_Week2']

# Sort the data by 'Revenue Loss' to find the most significant losses (in descending order)
comparison_data_sorted = comparison_data.sort_values(by='Revenue Loss', ascending=False)

# Filter for the top 50 products with the most significant revenue loss
top_50_revenue_loss_products = comparison_data_sorted.head(20)

# Drop duplicates based on 'Product Variant' to keep only one instance of each product
top_50_revenue_loss_products_unique = top_50_revenue_loss_products.drop_duplicates(subset='Product Variant')

# Display the filtered top 50 products with the most significant revenue loss from Week 1 to Week 2
print(top_50_revenue_loss_products_unique[['Product Variant', 'Revenue Loss']])







# In[156]:


import pandas as pd
import matplotlib.pyplot as plt

# Filter data for Week 1 and Week 2
week_1_data = df[df['week'] == 3]
week_2_data = df[df['week'] == 4]

# Aggregate sales data by 'Product Variant' for Week 1
week_1_aggregated = week_1_data.groupby('Product Variant', as_index=False)['Total'].sum()

# Aggregate sales data by 'Product Variant' for Week 2
week_2_aggregated = week_2_data.groupby('Product Variant', as_index=False)['Total'].sum()

# Merge the aggregated data for Week 1 and Week 2
comparison_data = pd.merge(
    week_1_aggregated, 
    week_2_aggregated, 
    on='Product Variant', 
    how='outer', 
    suffixes=('_Week1', '_Week2')
)

# Fill NaN values with 0 for products that might be missing in either week
comparison_data.fillna(0, inplace=True)

# Calculate the revenue gain between Week 1 and Week 2 (Sales in Week 2 - Sales in Week 1)
comparison_data['Revenue Gain'] = comparison_data['Total_Week2'] - comparison_data['Total_Week1']

# Filter products where the revenue gain is positive (i.e., Week 2 sales > Week 1 sales)
positive_revenue_gain = comparison_data[comparison_data['Revenue Gain'] > 0]

# Sort the data by 'Revenue Gain' to find the most significant gains (in descending order)
positive_revenue_gain_sorted = positive_revenue_gain.sort_values(by='Revenue Gain', ascending=False)

# Filter for the top 50 products with the most significant revenue gain
top_50_revenue_gain_products = positive_revenue_gain_sorted.head(20)

# Display the filtered top 50 products with the most significant revenue gain from Week 1 to Week 2
print(top_50_revenue_gain_products[['Product Variant', 'Revenue Gain']])


# In[157]:


# Step 1: Filter the data for week 3
week_3_df = df[df['week'] == 4]

# Step 2: Group by 'Company' and 'Order Date' to calculate total quantity per bill
grouped = (
    week_3_df.groupby(['Company', 'Order Date'])
    .agg({'Qty Ordered': 'sum'})  # Sum the quantities per bill
    .reset_index()
)

# Step 3: Categorize bills based on total quantity sold
grouped['WEEK 3'] = pd.cut(
    grouped['Qty Ordered'],
    bins=[0, 4, 9, float('inf')],
    labels=['1–4 Qty', '5–9 Qty', '10+ Qty'],
    right=True
)

# Step 4: Calculate total bills and categorize counts for each company
summary = (
    grouped.groupby(['Company', 'WEEK 3'])
    .size()  # Count the number of bills in each category
    .unstack(fill_value=0)  # Transform the categories into columns
    .reset_index()
)

# Step 5: Calculate total bills per company
summary['Total NoB'] = summary[['1–4 Qty', '5–9 Qty', '10+ Qty']].sum(axis=1)

# Step 6: Calculate percentages for each category
summary['1–4 Qty (%)'] = (summary['1–4 Qty'] / summary['Total NoB'] * 100).round(2)
summary['5–9 Qty (%)'] = (summary['5–9 Qty'] / summary['Total NoB'] * 100).round(2)
summary['10+ Qty (%)'] = (summary['10+ Qty'] / summary['Total NoB'] * 100).round(2)

# Step 7: Reorganize columns for clarity
summary = summary[['Company', 'Total NoB', '1–4 Qty (%)', '5–9 Qty (%)', '10+ Qty (%)']]

# Step 8: Display the summary
print(summary)




# In[158]:


# Step 1: Filter the data for week 3
week_3_df = df[df['week'] == 3]

# Step 2: Group by 'Company' and 'Order Date' to calculate total quantity per bill
grouped = (
    week_3_df.groupby(['Company', 'Order Date'])
    .agg({'Qty Ordered': 'sum'})  # Sum the quantities per bill
    .reset_index()
)

# Step 3: Categorize bills based on total quantity sold
grouped['WEEK 4'] = pd.cut(
    grouped['Qty Ordered'],
    bins=[0, 4, 9, float('inf')],
    labels=['1–4 Qty', '5–9 Qty', '10+ Qty'],
    right=True
)

# Step 4: Calculate total bills and categorize counts for each company
summary1 = (
    grouped.groupby(['Company', 'WEEK 4'])
    .size()  # Count the number of bills in each category
    .unstack(fill_value=0)  # Transform the categories into columns
    .reset_index()
)

# Step 5: Calculate total bills per company
summary1['Total NoB'] = summary1[['1–4 Qty', '5–9 Qty', '10+ Qty']].sum(axis=1)

# Step 6: Calculate percentages for each category
summary1['1–4 Qty (%)'] = (summary1['1–4 Qty'] / summary1['Total NoB'] * 100).round(2)
summary1['5–9 Qty (%)'] = (summary1['5–9 Qty'] / summary1['Total NoB'] * 100).round(2)
summary1['10+ Qty (%)'] = (summary1['10+ Qty'] / summary1['Total NoB'] * 100).round(2)

# Step 7: Reorganize columns for clarity
summary1 = summary1[['Company', 'Total NoB', '1–4 Qty (%)', '5–9 Qty (%)', '10+ Qty (%)']]

# Step 8: Display the summary
print(summary1)


# In[159]:


import pandas as pd

# Assuming df is your DataFrame and 'week' is already a column in your data

# Filter data for the 4th week
fourth_week_data = df[df['week'] == 4]

# Group by 'Product Variant' and aggregate 'Total' sales and 'Qty Ordered'
top_products_fourth_week = fourth_week_data.groupby('Product Variant').agg({'Total': 'sum', 'Qty Ordered': 'sum'}).sort_values(by='Total', ascending=False).head(20)

# Print the top 5 products based on total sales for the 4th week
print(top_products_fourth_week)


# Filter data for the 4th week and 'RJJ' company
rjj_data = df[(df['Company'] == 'RJJ') & (df['week'] == 4)]

# Group by 'Product Variant' and aggregate 'Total' sales and 'Qty Ordered'
top_products_rjj = (
    rjj_data.groupby('Product Variant')
    .agg({'Total': 'sum', 'Qty Ordered': 'sum'})
    .sort_values(by='Total', ascending=False)
    .head(20)
)

# Print the top 25 products based on total sales for the company 'RJJ'
print(top_products_rjj)


# In[160]:


from openpyxl import Workbook

file_path = "/Users/devnikhil/Downloads/juneweek2.xlsx"
wb = Workbook()

# Remove the default sheet created by openpyxl
default_sheet = wb.active
wb.remove(default_sheet)

# Get the list of unique branches
unique_branches = df['Company'].unique()

# Iterate through each branch and calculate metrics for week 6
for branch in unique_branches:
    # Filter the DataFrame for the current branch and week 6
    branch_data = df[(df['Company'] == branch) & (df['week'] == 4)]
    # Calculate the required metrics
    total_sales = branch_data['Total'].sum()
    total_quantity_sold = branch_data['Qty Ordered'].sum()
    unique_skus = branch_data['Product Variant'].nunique()
    average_order_value = total_sales / branch_data.drop_duplicates(subset=['Order Date']).shape[0] if branch_data.shape[0] > 0 else 0
    number_of_bills = branch_data.drop_duplicates(subset=['Order Date']).shape[0]
    skus_per_bill = total_quantity_sold/number_of_bills
    avg_daily_sales = total_sales / 7 if branch_data.shape[0] > 0 else 0
    # Create a new sheet for each branch
    ws = wb.create_sheet(title=f"{branch} All Dates Metrics")

    # Write the metrics to the Excel sheet in the specified order
    ws.append(["Metric", "Value"])
    ws.append(["Sales", total_sales])
    ws.append(["Avg Daily Sales", avg_daily_sales])
    ws.append(["AOV", average_order_value])
    ws.append(["NOB", number_of_bills])
    ws.append(["Qty Sold", total_quantity_sold])
    ws.append(["Unique SKUs", unique_skus])
    ws.append(["SKUs per bill", skus_per_bill])

    # Leave a blank row before the date-wise sales
    ws.append([])

    # Group by date and sum the sales across all dates
    date_wise_sales = branch_data.groupby(branch_data['Order Date'].dt.date)['Total'].sum()

    # Write the date-wise sales to the Excel sheet
    ws.append(["Date", "Sales"])
    for date, sales in date_wise_sales.items():
        ws.append([date, sales])


ws_top_sku = wb.create_sheet(title="Top SKU")
ws_top_sku.append(list(most_selling_sku_data.columns))  # Add column headers
for row in most_selling_sku_data.itertuples(index=False):
    ws_top_sku.append(row)  # Add data rows

# Create a new sheet for "Total Sales" data
ws_total_sales = wb.create_sheet(title="Total Sales")
ws_total_sales.append(list(branch_total_sales_df.columns))  # Add column headers
for row in branch_total_sales_df.itertuples(index=False):
    ws_total_sales.append(row)  # Add data rows

# Create a new sheet for the comparison data
ws_comparison = wb.create_sheet(title="Week 3 vs Week 4 Comparison")
ws_comparison.append(list(comparison_df.columns))
for row in comparison_df.itertuples(index=False):
    ws_comparison.append(row)

# Create a new sheet for the weekly summary data
ws_weekly_summary = wb.create_sheet(title="Weekly Summary")
ws_weekly_summary.append(list(weekly_summary_df.columns))
for row in weekly_summary_df.itertuples(index=False):
    ws_weekly_summary.append(row)

ws_revenue_loss = wb.create_sheet(title="Revenue Loss Top 20")
ws_revenue_loss.append(['Product Variant', 'Revenue Loss'])  # Adding headers for Revenue Loss
for row in top_50_revenue_loss_products_unique.itertuples(index=False, name=None):  # Use name=None for unnamed tuples
    ws_revenue_loss.append([row[0], row[3]])  # Access by position (0-based index)

# Create Revenue Gain sheet
ws_revenue_gain = wb.create_sheet(title="Revenue Gain Top 20")
ws_revenue_gain.append(['Product Variant', 'Revenue Gain'])  # Adding headers for Revenue Gain

for row in top_50_revenue_gain_products.itertuples(index=False, name=None):  # Use name=None for unnamed tuples
    ws_revenue_gain.append([row[0], row[3]])  # Correctly reference column 4 ('Revenue Gain')

ws_week_4 = wb.create_sheet("Week 4 Summary")
ws_week_4.append(list(summary.columns))
for row in summary.itertuples(index=False):
    ws_week_4.append(row)
ws_week_4.append(list(summary1.columns))
for row in summary1.itertuples(index=False):
    ws_week_4.append(row)

ws_week4 = wb.create_sheet(title="Top 20 Products - Week 4")
ws_week4.append(['Product Variant', 'Total Sales', 'Qty Ordered'])  # Adding headers
for row in top_products_fourth_week.itertuples(name=None):  # Iterate without index
    ws_week4.append([row[0], row[1], row[2]])  # Product Variant, Total Sales, Qty Ordered

# Create "Top 25 Products - Week 4 RJJ" sheet
ws_rjj = wb.create_sheet(title="Top 25 Products - Week 4 RJJ")
ws_rjj.append(['Product Variant', 'Total Sales', 'Qty Ordered'])  # Adding headers
for row in top_products_rjj.itertuples(name=None):  # Iterate without index
    ws_rjj.append([row[0], row[1], row[2]])  # Product Variant, Total Sales, Qty Ordered

ws_summary = wb.create_sheet(title="Company Sales Summary - Week 4")
ws_summary.append([
    'Company', 
    'Total Sales', 
    'Total Normal Sales', 
    'Total 50% Discount Sales', 
    'Total Normal Sales %', 
    'Total 50% Discount Sales %'
])
for row in company_sales_summary_week_4.itertuples(index=False, name=None):  # Use name=None for unnamed tuples
    ws_summary.append(row)
wb.save(file_path)

print(f"Data successfully added to {file_path}.")


# In[ ]:





# In[ ]:





# In[ ]:




